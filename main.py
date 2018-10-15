#quandl.ApiConfig.api_key = "HWWiy9cYQYxJhaoCi1ac"
#data = quandl.get_table('MER/F1', compnumber="39102", paginate=True)
#import quandl
#https://finviz.com/screener.ashx
from tkinter import *
from tkinter import ttk
from bs4 import BeautifulSoup
import requests
import json
import time
import xlsxwriter
import os
from openpyxl import load_workbook
from openpyxl import Workbook

window = Tk()
actions = []

def organize(vHeader):
	if vHeader == '311':
		minusColumn = 0
		moreRow = 1
		moreColumn = 1
		columnRange = 71
		rowRange = 1

		beforeCol = 71
		beforeRow = 1

	if vHeader == '171':
		minusColumn = 300
		moreRow = 1
		moreColumn = 1
		columnRange = 66
		rowRange = 1

		beforeCol = 66
		beforeRow = 1

	if vHeader == '141':
		minusColumn = 320
		moreRow = 1
		moreColumn = 1
		columnRange = 51
		rowRange = 1

		beforeCol = 51
		beforeRow = 1

	if vHeader == '131':
		minusColumn = 300
		moreRow = 1
		moreColumn = 1
		columnRange = 36
		rowRange = 1

		beforeCol = 36
		beforeRow = 1

	if vHeader == '161':
		minusColumn = 360
		moreRow = 1
		moreColumn = 0
		columnRange = 19
		rowRange = 1

		beforeCol = 19
		beforeRow = 1
		
	if vHeader == '121':
		minusColumn = 360
		moreRow = 2
		moreColumn = 1
		rowRange = 1
		columnRange = 1
		beforeCol = 1
		beforeRow = 1

	# Create Excel File
	if vHeader == '121':
		wb = Workbook()
		fecha = time.strftime("%d-%m-%y")
		filepathNew = "./acciones_"+str(fecha)+".xlsx"
		wb.save(filepathNew)
	# Load Excel File
	filepath = "./acciones"+vHeader+".xlsx"
	wb = load_workbook(filepath)
	sheet = wb.active

	
	filepathNewFile = "./acciones_fechadehoy.xlsx"
	newFile = load_workbook(filepathNewFile)		
	sheetnewFile = newFile.active
	
	max_column = (sheet.max_column - minusColumn)
	max_row = sheet.max_row

	for i in range(1, max_row + moreRow):
		columnRange = beforeCol
		for j in range(1, max_column + moreColumn):
			oldCell = sheet.cell(row = i, column = j).value
			if vHeader != '121':
				sheetnewFile.cell(row = rowRange, column = columnRange).value = oldCell
				columnRange = columnRange + 1
			else: 
				sheetnewFile.cell(row = i, column = j).value = oldCell
		rowRange = rowRange + 1
	newFile.save(filepathNewFile)

def organizeTables():
	vHeaders = ['121', '161', '131', '141', '171', '311']
	for vHeader in vHeaders:
		organize(vHeader)

def downloadTables(vHeader):
	workbook = xlsxwriter.Workbook('acciones'+vHeader+'.xlsx')
	worksheet = workbook.add_worksheet()
	worksheet.set_column('A:A', 20)

	if vHeader == '311':
		i = 0
		y = 0

		while i < 7531:
			if i == 0:
				req = requests.get('https://finviz.com/screener.ashx?v=311')
			req = requests.get('https://finviz.com/screener.ashx?v=311&r=' + str(i))

			soup = BeautifulSoup(req.text, 'html.parser')
			screener = soup.findAll('div', {'id': 'screener-content'})
			tables = screener[0].findAll('table', {'class': 'snapshot-table'})
			s = 0
			for table in tables:
				rows = table.findAll('tr', {'class': 'table-light2-row'})
				for row in rows:
					cells = row.findAll('td', {'class': 'snapshot-td'})
					x = 0
					for cell in cells:
						text = cell.get_text()
						worksheet.write( y, x, text)
						x = x + 1
					y = y + 1

				secondTables = screener[0].findAll('table', {'class': 'snapshot-table2'})
				table = secondTables[s]
				rows = table.findAll('tr', {'class': 'table-dark-row'})
				for row in rows:
					cells = row.findAll('td', {'class': 'snapshot-td2'})
					x = 0
					for cell in cells:
						text = cell.get_text()
						worksheet.write( y, x, text)
						x = x + 1
					y = y + 1
				worksheet.write(y, x,'\n')
				s = s + 1
			if i == 0:
				i = i + 11
			i = i + 10
		workbook.close()
		

	i = 0
	y = 0
	while i < 7521:
		if i == 0:
			req = requests.get('https://finviz.com/screener.ashx?v=' +str(vHeader))

		req = requests.get('https://finviz.com/screener.ashx?v=' +str(vHeader)+'&r='+str(i))
		soup = BeautifulSoup(req.text, 'html.parser')
		screener = soup.findAll('div', {'id': 'screener-content'})	
		table = screener[0].findAll('table', {'bgcolor': '#d3d3d3'})
		rows = table[0].findAll('tr')
		for row in rows:
			cells = row.findAll('td')
			x = 0
			for cell in cells:
				text = cell.get_text()
				worksheet.write(y, x, text)
				x = x + 1
			y = y + 1

		if i == 0:
			i = i + 21
		i = i + 20

	workbook.close()
	if vHeader == '121':
		organizeTables()

def downloadAllTables():
	vHeaders = ['311', '171', '141', '131', '161', '121']
	for vHeader in vHeaders:
		downloadTables(vHeader)
	
def downloadImg():

	os.mkdir('graficas')

	header = 0
	while header < 7441:
		if header == 0:
			req = requests.get('https://finviz.com/screener.ashx')
			soup = BeautifulSoup(req.text, 'html.parser')

			acciones = soup.findAll('tr', {'class': 'table-dark-row-cp'})
			for accion in acciones:
				celdas = accion.findAll('td', {'class': 'screener-body-table-nw'})
				nombreAccion = celdas[1].get_text()
				actions.append(nombreAccion)

			accionesW = soup.findAll('tr', {'class': 'table-light-row-cp'})
			for accionW in accionesW:
				celdasW = accionW.findAll('td', {'class': 'screener-body-table-nw'})
				nombreAccionW = celdasW[1].get_text()
				actions.append(nombreAccionW)

			for action in actions:
				page = requests.get('https://finviz.com/chart.ashx?t='+action+'&ta=1&ty=c&p=d&s=l')
				img = page.content	
				with open('graficas/'+action+'.png', 'wb') as handler:
					handler.write(img)
			header = header + 21	
		
		req = requests.get('https://finviz.com/screener.ashx?v=111&r='+str(header))
		soup = BeautifulSoup(req.text, 'html.parser')

		acciones = soup.findAll('tr', {'class': 'table-dark-row-cp'})
		for accion in acciones:
			celdas = accion.findAll('td', {'class': 'screener-body-table-nw'})
			nombreAccion = celdas[1].get_text()
			actions.append(nombreAccion)

		accionesW = soup.findAll('tr', {'class': 'table-light-row-cp'})
		for accionW in accionesW:
			celdasW = accionW.findAll('td', {'class': 'screener-body-table-nw'})
			nombreAccionW = celdasW[1].get_text()
			actions.append(nombreAccionW)

		for action in actions:
			page = requests.get('https://finviz.com/chart.ashx?t='+action+'&ta=1&ty=c&p=d&s=l')
			img = page.content	
			with open('graficas/'+action+'.png', 'wb') as handler:
				handler.write(img)	

		header = header + 21
def createFile():
	fecha = time.strftime("%d-%m-%y")
	file = open('acciones_'+str(fecha)+'.csv', 'w')
	file.write('No.' + ',' + 'Ticker' + ',' + 'Company' + ',' + 'Sector' + ',' + 'Industry' + ',' + 'Country' + ',' + 'Market Cap' + ',' + 'P/E' + ',' + 'Price' + ',' + 'Change' + ',' + 'Volume' + '\n')
	header = 0
	while header < 7441:
		if header == 0:
			url = 'https://finviz.com/screener.ashx'
			r = requests.get(url)
			soup = BeautifulSoup(r.text, 'html.parser')
			actions = soup.findAll('tr', {'class': 'table-dark-row-cp'})

			for action in actions:
				cells = action.findAll('td', {'class': 'screener-body-table-nw'})
				for cell in cells:
					file.write(cell.get_text() + ',')
				file.write('\n')

			actionsW = soup.findAll('tr', {'class': 'table-light-row-cp'})

			for actionW in actionsW:
				cellsW = actionW.findAll('td', {'class': 'screener-body-table-nw'})
				for cellW in cellsW:
					file.write(cellW.get_text() + ',')
				file.write('\n')

			header = header + 21
		

		url = 'https://finviz.com/screener.ashx?v=111&r='+str(header)
		r = requests.get(url)
		soup = BeautifulSoup(r.text, 'html.parser')
		actions = soup.findAll('tr', {'class': 'table-dark-row-cp'})

		for action in actions:
			cells = action.findAll('td', {'class': 'screener-body-table-nw'})
			for cell in cells:
				file.write(cell.get_text() + ',')
			file.write('\n')
		
		actionsW = soup.findAll('tr', {'class': 'table-light-row-cp'})

		for actionW in actionsW:
			cellsW = actionW.findAll('td', {'class': 'screener-body-table-nw'})
			for cellW in cellsW:
				file.write(cellW.get_text() + ',')
			file.write('\n')
		header = header + 20
	file.close()

	workbook = xlsxwriter.Workbook('acciones.xlsx')
	worksheet = workbook.add_worksheet()

	worksheet.set_column('A:A', 20)
	fecha = time.strftime("%d-%m-%y")
	csv = open('acciones_'+str(fecha)+'.csv', 'r')
	csvLines = csv.readlines()

	y = 0
	
	for line in csvLines:
		lineArray = str(line).split(',')
		x = 0
		for cell in lineArray:
			worksheet.write(y, x, cell)
			x = x + 1
		y = y + 1
	#print(str(csvLines))

	workbook.close()
	return main()
	
def searchAction(url, value):
	#key =  'IX09TC435VGY2C5F'
	#r = requests.get('https://www.alphavantage.co/query?function=TIME_SERIES_DAILY&symbol='+value+'&apikey='+key)
	#print(r.text)
	r = requests.get(url)
	soup = BeautifulSoup(r.text, 'html.parser')
	actions = soup.findAll('tr', {'class': 'table-dark-row-cp'})

	for action in actions:
		cells = action.findAll('td', {'class': 'screener-body-table-nw'})
		if cells[1].get_text() == value:
			newWindow = Tk()
			newFrame = ttk.Frame(newWindow, padding = '100 100 40 20', borderwidth = 20, relief = 'sunken')
			ttk.Label(newWindow, text= 'Acción Requerida: ' + str(cells[1].get_text()), foreground = 'purple').place(x=50, y=100)
			ttk.Label(newWindow, text= 'P/E: ' + str(cells[7].get_text()), foreground  = 'green').place(x=50, y=120)
			ttk.Label(newWindow, text= 'Price: ' + str(cells[8].get_text()), foreground = 'red').place(x=50, y=140)
			ttk.Label(newWindow, text= 'Change: ' + str(cells[9].get_text()), foreground = 'brown').place(x=50, y=160)
			ttk.Label(newWindow, text= 'Change: ' + str(cells[10].get_text()), foreground = 'brown').place(x=50, y=180)
			return main()

	actionsW = soup.findAll('tr', {'class': 'table-light-row-cp'})
	for actionW in actionsW:
		cellsW = actionW.findAll('td', {'class': 'screener-body-table-nw'})
		if cellsW[1].get_text() == value:
			newWindow = Tk()
			newFrame = ttk.Frame(newWindow, padding = '100 100 40 20', borderwidth = 20, relief = 'sunken')
			ttk.Label(newWindow, text= 'Acción Requerida: ' + str(cellsW[1].get_text()), foreground = 'purple').place(x=50, y=100)
			ttk.Label(newWindow, text= 'P/E: ' + str(cellsW[7].get_text()), foreground  = 'green').place(x=50, y=120)
			ttk.Label(newWindow, text= 'Price: ' + str(cellsW[8].get_text()), foreground = 'red').place(x=50, y=140)
			ttk.Label(newWindow, text= 'Change: ' + str(cellsW[9].get_text()), foreground = 'brown').place(x=50, y=160)
			ttk.Label(newWindow, text= 'Change: ' + str(cellsW[10].get_text()), foreground = 'brown').place(x=50, y=180)
			return main()

def searchLoop(value):
	header = 0
	while header < 7441:
		if header == 0:
			url = 'https://finviz.com/screener.ashx'
			searchAction(url, value)
			header = header + 21

		url = 'https://finviz.com/screener.ashx?v=111&r='+str(header)
		searchAction(url, value)
		header = header + 20

def main():
	window.geometry("300x200")
	frame = ttk.Frame(window, padding = '100 100 40 20', borderwidth = 20, relief = 'sunken')
	ttk.Label(window, text= "Ingrese la Acción").place(x=110, y=9)
	value = StringVar()
	entry = ttk.Entry(window, width=15, textvariable = value).place(x=110, y=35)
	ttk.Button(window, text='Buscar', command = lambda: searchLoop(value.get())).place(x = 120, y = 60)
	ttk.Button(window, text='Crear Archivo', command = createFile).place(x = 70, y = 90)
	ttk.Button(window, text='Descargar Gráficas', command = downloadImg).place(x = 160, y = 90)
	ttk.Button(window, text='Descargar Tablas', command = downloadAllTables).place(x = 100, y = 130)
	window.mainloop()

if __name__ == '__main__':
	main()
	